from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import datetime, timedelta
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from api.database import get_db
from api.models import Order, OrderItem, OrderStatus
from api.dependencies import verify_admin_token

router = APIRouter()

@router.get("/csv")
async def export_csv(
    status: str = None,
    date_from: str = None,
    date_to: str = None,
    db: AsyncSession = Depends(get_db),
    admin=Depends(verify_admin_token)
):
    query = select(Order).order_by(Order.created_at.desc())
    if status:
        query = query.where(Order.status == status)
    if date_from:
        query = query.where(Order.created_at >= datetime.fromisoformat(date_from))
    if date_to:
        query = query.where(Order.created_at <= datetime.fromisoformat(date_to))

    result = await db.execute(query)
    orders = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "User ID", "Status", "Total", "Discount", "Final", "Address", "Phone", "Date", "Items"])

    for order in orders:
        items_str = "; ".join([f"{item.product.name} x{item.quantity}" for item in order.items])
        writer.writerow([
            order.id, order.user_id, order.status.value,
            float(order.total_amount), float(order.discount_amount or 0), float(order.final_amount),
            order.shipping_address, order.phone, order.created_at.isoformat(), items_str
        ])

    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
    )

@router.get("/excel")
async def export_excel(
    status: str = None,
    date_from: str = None,
    date_to: str = None,
    db: AsyncSession = Depends(get_db),
    admin=Depends(verify_admin_token)
):
    query = select(Order).order_by(Order.created_at.desc())
    if status:
        query = query.where(Order.status == status)
    if date_from:
        query = query.where(Order.created_at >= datetime.fromisoformat(date_from))
    if date_to:
        query = query.where(Order.created_at <= datetime.fromisoformat(date_to))

    result = await db.execute(query)
    orders = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="f4a261", end_color="f4a261", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    status_colors = {"pending": "fef3c7", "processing": "dbeafe", "shipped": "e0e7ff", "delivered": "d1fae5", "cancelled": "fee2e2"}

    headers = ["ID", "Status", "Customer", "Phone", "Address", "Items", "Total", "Discount", "Final", "Date"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row, order in enumerate(orders, 2):
        items_text = "\n".join([f"• {item.product.name} x{item.quantity} = {float(item.price) * item.quantity:.0f}₽" for item in order.items])

        ws.cell(row=row, column=1, value=order.id).border = thin_border
        status_cell = ws.cell(row=row, column=2, value=order.status.value)
        status_cell.fill = PatternFill(start_color=status_colors.get(order.status.value, "ffffff"), fill_type="solid")
        status_cell.border = thin_border
        ws.cell(row=row, column=3, value=order.user_id).border = thin_border
        ws.cell(row=row, column=4, value=order.phone).border = thin_border
        ws.cell(row=row, column=5, value=order.shipping_address).border = thin_border
        items_cell = ws.cell(row=row, column=6, value=items_text)
        items_cell.alignment = Alignment(wrap_text=True, vertical="top")
        items_cell.border = thin_border
        ws.cell(row=row, column=7, value=float(order.total_amount)).border = thin_border
        ws.cell(row=row, column=8, value=float(order.discount_amount or 0)).border = thin_border
        ws.cell(row=row, column=9, value=float(order.final_amount)).border = thin_border
        ws.cell(row=row, column=10, value=order.created_at.strftime("%d.%m.%Y %H:%M")).border = thin_border

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = [8, 15, 12, 18, 35, 40, 12, 12, 12, 18][col-1]

    for row in range(2, len(orders) + 2):
        ws.row_dimensions[row].height = 60

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:J{len(orders) + 1}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )
